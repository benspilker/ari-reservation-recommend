###############################################################################
# Azure Reservation Recommendation Tool: Expanding on AzureResourceInventory
###############################################################################


# Standard library imports
import os
import sys
import subprocess
import glob
import json
import time
import threading


###############################################################################
# DEPENDENCY MANAGEMENT
###############################################################################

def install_and_import(package_name, import_name=None):
    """
    Automatically install and import a package if it's not available.
    
    Args:
        package_name: The name of the package to install via pip
        import_name: The name to use for importing (if different from package_name)
    """
    if import_name is None:
        import_name = package_name
    
    try:
        __import__(import_name)
        print(f"✓ {package_name} is already installed")
    except ImportError:
        print(f"⚠️ {package_name} not found. Installing...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])
            print(f"✅ Successfully installed {package_name}")
        except subprocess.CalledProcessError as e:
            print(f"❌ Failed to install {package_name}: {e}")
            sys.exit(1)


def check_and_install_dependencies():
    """Check and install all required dependencies"""
    print("\n" + "="*80)
    print("CHECKING DEPENDENCIES")
    print("="*80 + "\n")
    
    required_packages = [
        ("pandas", "pandas"),
        ("openpyxl", "openpyxl"),
        ("requests", "requests"),
        ("selenium", "selenium"),
    ]
    
    for package_name, import_name in required_packages:
        install_and_import(package_name, import_name)
    
    print("\n✅ All dependencies are installed!\n")


# Check dependencies before importing
check_and_install_dependencies()

# Third-party imports (after dependency check)
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import requests


###############################################################################
# SECTION 1: VM SUGGESTIONS AND RECOMMENDATIONS
###############################################################################

def generate_vm_recommendations():
    """Generate VM reservation recommendations from Azure Resource Inventory"""
    print("\n" + "="*80)
    print("SECTION 1: GENERATING VM RESERVATION RECOMMENDATIONS")
    print("="*80 + "\n")
    
    # Define a variable for the savings threshold
    savings_threshold = 10

    # === Auto-detect Azure Resource Inventory file ===
    matching_files = glob.glob("AzureResourceInventory_Report_*.xlsx")

    if not matching_files:
        raise FileNotFoundError("No file found matching AzureResourceInventory_Report_*.xlsx")

    # If multiple files, pick the newest by modified date
    file_path = max(matching_files, key=os.path.getmtime)
    print(f"Processing file: {os.path.basename(file_path)}")

    # === Load Workbook ===
    advisor_df = pd.read_excel(file_path, sheet_name="Advisor")
    vm_df = pd.read_excel(file_path, sheet_name="Virtual Machines")

    # === Filter only running VMs ===
    if "Power State" in vm_df.columns:
        initial_vm_count = len(vm_df)
        vm_df = vm_df[vm_df["Power State"].astype(str).str.strip().eq("VM running")]
        print(f"Filtered Virtual Machines: {len(vm_df)} of {initial_vm_count} are running (Power State = 'VM running')")
    else:
        print("Warning: 'Power State' column not found — including all VMs.")

    # === Filter Advisor ===
    advisor_filtered = advisor_df[
        (advisor_df["Category"] == "Cost")
        & (advisor_df["Impact"] == "High")
        & (advisor_df["Description"].str.contains("reserved instance", case=False, na=False))
    ]

    print(f"Filtered Advisor rows: {len(advisor_filtered)}")

    # === Total Quantity in Advisor ===
    total_quantity = advisor_filtered["Quantity"].sum()
    print(f"Total quantity of VMs in Advisor sheet: {int(total_quantity)}")

    # === Build VM Pool with Tag Awareness (ApplicationName + CostCenter fallback) ===
    vm_pool = {}
    grouped_vms = vm_df.groupby("VM Name")

    for vm_name, group in grouped_vms:
        vm_size = group["VM Size"].iloc[0] if "VM Size" in group else "Unknown"
        os_type = group["OS Type"].iloc[0] if "OS Type" in group else "Unknown"
        os_name = group["OS Name"].iloc[0] if "OS Name" in group else "Unknown"
        location = group["Location"].iloc[0] if "Location" in group else "Unknown"

        # Default tag value
        tag_value = "Unknown"

        # Prefer ApplicationName, else fallback to CostCenter
        if "Tag Name" in group and "Tag Value" in group:
            app_row = group[group["Tag Name"] == "ApplicationName"]
            if not app_row.empty:
                tag_value = str(app_row["Tag Value"].iloc[0]).strip()
            else:
                cost_row = group[group["Tag Name"].isin(["CostCenter", "Cost Center"])]
                if not cost_row.empty:
                    tag_value = str(cost_row["Tag Value"].iloc[0]).strip()

        key = (str(vm_size).strip(), str(location).lower().strip())
        vm_entry = {
            "VM Name": vm_name,
            "VM Size": vm_size,
            "OS": os_type,
            "OS Name": os_name,
            "Region": location,
            "Tags": tag_value,
        }
        vm_pool.setdefault(key, []).append(vm_entry)

    # === Build Recommendations ===
    recommendations = []
    seen_vms = set()

    for _, adv in advisor_filtered.iterrows():
        key = (str(adv["SKU"]).strip(), str(adv["Savings Region"]).lower().strip())
        quantity = int(adv["Quantity"])
        annual_savings = float(adv["Annual Savings"])
        pool = vm_pool.get(key, [])

        selected = 0
        for vm in pool:
            if vm["VM Name"] in seen_vms:
                continue
            recommendations.append(
                {
                    "Subscription": adv["Name"],
                    "Recommendations": [
                        {
                            "VM Name": vm["VM Name"],
                            "VM Size": vm["VM Size"],
                            "SKU": adv["SKU"],
                            "Recommendation": adv["Description"],
                            "Annual Savings": annual_savings,
                            "Impact": adv["Impact"],
                            "Region": adv["Savings Region"],
                            "OS": vm["OS"],
                            "OS Name": vm["OS Name"],
                            "Tags": vm["Tags"],
                        }
                    ],
                }
            )
            seen_vms.add(vm["VM Name"])
            selected += 1
            if selected >= quantity:
                break

    # === Sort Recommendations by Annual Savings (High → Low) ===
    recommendations.sort(
        key=lambda rec: rec["Recommendations"][0]["Annual Savings"], reverse=True
    )

    # === Filter High-Impact Recommendations (≥ savings_threshold) ===
    impact_recs = [
        rec for rec in recommendations if rec["Recommendations"][0]["Annual Savings"] >= savings_threshold
    ]

    # If less than 20 high-impact recommendations, add more from the lower range (>= $1)
    if len(impact_recs) < 20:
        impact_recs += [
            rec for rec in recommendations if rec["Recommendations"][0]["Annual Savings"] > 1
            and rec not in impact_recs
        ][: 20 - len(impact_recs)]

    # === Save output.json (high-impact recommendations only) ===
    with open("output.json", "w") as f:
        json.dump(impact_recs, f, indent=4)

    # === Summary ===
    unique_skus = set([rec["Recommendations"][0]["SKU"] for rec in recommendations])
    os_counts = {}
    for rec in recommendations:
        os_type = rec["Recommendations"][0]["OS"]
        os_counts[os_type] = os_counts.get(os_type, 0) + 1

    print("\n=== Summary Report ===")
    print(f"Unique SKUs: {len(unique_skus)}")
    print(f"Total VMs: {len(recommendations)}")
    print(f"OS Distribution: {os_counts}")

    # === Impact Summary ===
    total_savings = sum(rec["Recommendations"][0]["Annual Savings"] for rec in impact_recs)
    impact_summary = {
        "Total VMs": len(impact_recs),
        "Savings (USD) According to ARI": round(total_savings, 2),
        "Average Savings per VM (USD)": round(total_savings / len(impact_recs), 2)
        if impact_recs
        else 0,
    }
    print("\n=== Impact Summary ===")
    print(json.dumps(impact_summary, indent=4))

    # === Create inputs.json for automation ===
    input_json = [
        {
            "Region": rec["Recommendations"][0]["Region"],
            "SKU": rec["Recommendations"][0]["SKU"],
            "OS": rec["Recommendations"][0]["OS"],
            "OS Name": rec["Recommendations"][0]["OS Name"],
            "VM Name": rec["Recommendations"][0]["VM Name"],
            "Tags": rec["Recommendations"][0]["Tags"],
        }
        for rec in impact_recs
    ]

    with open("inputs.json", "w") as f:
        json.dump(input_json, f, indent=4)
    
    print(f"\n✅ Created output.json with {len(impact_recs)} recommendations")
    print(f"✅ Created inputs.json with {len(input_json)} entries")


###############################################################################
# SECTION 2: PRICING ANALYSIS AND SPREADSHEET GENERATION
###############################################################################

def get_prices(sku, region):
    """Fetch prices from Azure Retail Prices API for a given SKU and region"""
    base_url = "https://prices.azure.com/api/retail/prices"
    filter_str = f"$filter=serviceName eq 'Virtual Machines' and armSkuName eq '{sku}' and armRegionName eq '{region}'"
    url = f"{base_url}?{filter_str}"

    all_items = []
    while url:
        r = requests.get(url)
        try:
            data = r.json()
        except json.decoder.JSONDecodeError:
            print(f"⚠️ Failed to decode JSON from {url}")
            break
        all_items.extend(data.get('Items', []))
        url = data.get('NextPageLink')
    return all_items


def matches_os(item, os_filter):
    """OS filter: Windows/Linux; always exclude Spot and Low Priority"""
    name = item.get("productName", "").lower()
    meter = item.get("meterName", "").lower()

    if "spot" in meter or "low priority" in meter:
        return False

    if not os_filter:
        return True

    os_filter = os_filter.lower()
    if os_filter == "windows":
        return "windows" in name or "windows" in meter
    elif os_filter == "linux":
        return "windows" not in name and "windows" not in meter
    return False


def build_azure_pricing(inputs):
    """Build Azure pricing data from API - runs in parallel thread"""
    estimate_rows = []
    price_cache = {}
    unique_skus_regions = set()

    for row in inputs:
        region = row["Region"]
        sku = row["SKU"]
        os_type = row.get("OS", "")
        os_name = row.get("OS Name", "")
        hostname = row.get("VM Name", "")
        tags = row.get("Tags", "")

        # Track unique SKU-region pairs for Windows
        if os_type.lower() == "windows":
            sanitized_sku = sku.replace("Standard_", "").lower().replace("_", "-")
            sanitized_sku_region = f"{sanitized_sku}_{region.lower()}"
            unique_skus_regions.add(sanitized_sku_region)
        else:
            sanitized_sku = sku.lower().replace("_", "-")
            sanitized_sku_region = f"{sanitized_sku}_{region.lower()}"

        print(f"\n🔍 Fetching {sku} in {region} ({os_type})")

        cache_key = (sku, region)
        if cache_key in price_cache:
            prices = price_cache[cache_key]
        else:
            prices = get_prices(sku, region)
            price_cache[cache_key] = prices

        if not prices:
            print(f"⚠️ No pricing data returned for {sku} in {region}")
            continue

        # === PAYG ===
        payg = [p for p in prices if p.get("type") == "Consumption" and matches_os(p, os_type)]
        if os_type.lower() == "linux":
            payg = sorted(payg, key=lambda x: x.get("unitPrice", 0))[:1]

        for p in payg:
            meter_name = p.get("meterName", "")
            price = p.get("unitPrice", 0)
            monthly_cost = round(price * 730, 2)
            estimate_rows.append({
                "Service category": "Compute",
                "Service type": "Virtual Machines",
                "VM Name": hostname,
                "Tags": tags,
                "Region": region,
                "OS": os_type,
                "OS Name": os_name,
                "SKU": sanitized_sku,
                "Description": f"1 {meter_name} ({sku}), {os_type}, Pay-as-you-go",
                "Estimated monthly cost": f"${monthly_cost:,.2f}"
            })

        # === Reservations ===
        if os_type.lower() == "windows":
            reservations = [
                p for p in prices
                if p.get("type") == "Reservation"
                and not any(term in (p.get("productName", "") + p.get("meterName", "")).lower() for term in ["ahb", "hybrid"])]

            if not reservations:
                print(f"⚠️ No clear Windows license-included reservation found for {sku}, using generic reservations")
                reservations = [p for p in prices if p.get("type") == "Reservation"]

            by_term = {}
            for p in reservations:
                term = p.get("reservationTerm", "")
                if term not in by_term or p.get("unitPrice", 0) > by_term[term].get("unitPrice", 0):
                    by_term[term] = p
            reservations = list(by_term.values())
        else:
            reservations = [p for p in prices if p.get("type") == "Reservation" and matches_os(p, os_type)]

        for p in reservations:
            meter_name = p.get("meterName", "")
            term = p.get("reservationTerm", "")
            total_price = p.get("unitPrice", 0)
            months = 12 if "1" in term else 36
            monthly_cost = round(total_price / months, 2)

            estimate_rows.append({
                "Service category": "Compute",
                "Service type": "Virtual Machines",
                "VM Name": hostname,
                "Tags": tags,
                "Region": region,
                "OS": os_type,
                "OS Name": os_name,
                "SKU": sanitized_sku,
                "Description": f"1 {meter_name} ({sku}) ({term}), {os_type} Reservation",
                "Estimated monthly cost": f"${monthly_cost:,.2f}"
            })

    return estimate_rows, unique_skus_regions


def scrape_windows_pricing(unique_skus_regions):
    """Scrape Windows pricing from vantage.sh - runs in parallel thread"""
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    driver = webdriver.Chrome(options=chrome_options)

    azure_windows_pricing_data = {}
    sku_region_pairs = list(unique_skus_regions)
    total_pairs = len(sku_region_pairs)

    for idx, sku_region in enumerate(sku_region_pairs, start=1):
        sku, region = sku_region.split('_')
        print(f"Processing {idx}/{total_pairs}: SKU = {sku}, Region = {region}...")

        url = f"https://instances.vantage.sh/azure/vm/{sku}?currency=USD&platform=windows&duration=monthly&pricingType=Standard.allUpfront&region={region}"
        driver.get(url)
        time.sleep(1)

        try:
            section = driver.find_element(By.CSS_SELECTOR, "section.mb-4")
            pricing_elements = section.find_elements(By.CSS_SELECTOR, "p.font-bold")
            pricing_data = {}

            for element in pricing_elements:
                pricing_text = element.text.strip().replace("\n", " ").split(" ")[0]
                pricing_data[pricing_text] = element.find_element(By.XPATH, "..").text.strip()

            azure_windows_pricing_data[sku_region] = pricing_data

        except Exception as e:
            print(f"⚠️ Error processing {sku_region}: {e}")

    driver.quit()

    # Clean and filter the pricing data
    filtered_pricing = {}
    for sku_region, prices in azure_windows_pricing_data.items():
        filtered_prices = {}
        for price, description in prices.items():
            description_clean = description.replace("\n", " ")
            if 'on demand' in description_clean.lower():
                filtered_prices["On Demand"] = price
            elif '1-year reserved' in description_clean.lower():
                filtered_prices["1-Year Reserved"] = price
            elif '3-year reserved' in description_clean.lower():
                filtered_prices["3-Year Reserved"] = price
        filtered_pricing[sku_region] = filtered_prices

    return filtered_pricing


def build_final_dataframes(estimate_rows, windows_pricing_data):
    """Build final sorted and formatted DataFrames"""
    # Create initial DataFrame
    df = pd.DataFrame(estimate_rows)
    
    # Extract pricing for sorting
    vm_pricing = {}
    for _, row in df.iterrows():
        if pd.notna(row['VM Name']) and 'Pay-as-you-go' in str(row.get('Description', '')):
            cost_str = str(row['Estimated monthly cost']).replace('$', '').replace(',', '')
            try:
                vm_pricing[row['VM Name']] = float(cost_str)
            except:
                vm_pricing[row['VM Name']] = 0
    
    # Sort VMs by pay-as-you-go cost
    sorted_vms = sorted(vm_pricing.items(), key=lambda x: x[1], reverse=True)
    sorted_vm_names = [vm[0] for vm in sorted_vms]
    
    # Rebuild dataframe in sorted order with blank rows
    sorted_rows = []
    for vm_name in sorted_vm_names:
        vm_rows = df[df['VM Name'] == vm_name]
        for _, row in vm_rows.iterrows():
            sorted_rows.append(row.to_dict())
        # Add blank row
        sorted_rows.append({col: '' for col in df.columns})
    
    # Calculate totals
    df_sorted = pd.DataFrame(sorted_rows)
    df_sorted['Estimated monthly cost'] = df_sorted['Estimated monthly cost'].replace('', None)
    
    # Extract numeric values for totals
    df_sorted['cost_numeric'] = df_sorted.apply(lambda row: 
        float(str(row['Estimated monthly cost']).replace('$', '').replace(',', '')) 
        if pd.notna(row['Estimated monthly cost']) and row['Estimated monthly cost'] != '' 
        else 0, axis=1)
    
    total_payg = df_sorted[df_sorted['Description'].str.contains('Pay-as-you-go', case=False, na=False)]['cost_numeric'].sum()
    total_1yr = df_sorted[df_sorted['Description'].str.contains('1 Year', case=False, na=False)]['cost_numeric'].sum()
    total_3yr = df_sorted[df_sorted['Description'].str.contains('3 Years', case=False, na=False)]['cost_numeric'].sum()
    
    # Add totals rows
    totals_rows = [
        {'Service category': 'Total', 'Service type': '', 'VM Name': '', 'Tags': '', 'Region': '', 
         'OS': '', 'OS Name': '', 'SKU': '', 'Description': 'Total Monthly Pay-as-you-go', 
         'Estimated monthly cost': f'${total_payg:,.2f}'},
        {'Service category': '', 'Service type': '', 'VM Name': '', 'Tags': '', 'Region': '', 
         'OS': '', 'OS Name': '', 'SKU': '', 'Description': 'Total 1 Year Reservations (Billed monthly)', 
         'Estimated monthly cost': f'${total_1yr:,.2f}'},
        {'Service category': '', 'Service type': '', 'VM Name': '', 'Tags': '', 'Region': '', 
         'OS': '', 'OS Name': '', 'SKU': '', 'Description': 'Total 3 Year Reservations (Billed monthly)', 
         'Estimated monthly cost': f'${total_3yr:,.2f}'},
    ]
    
    df_sorted = df_sorted.drop('cost_numeric', axis=1)
    df_final = pd.concat([df_sorted, pd.DataFrame(totals_rows)], ignore_index=True)
    
    # Build azure_savings_estimate with Windows pricing updates
    df_savings = df_final.copy()
    
    for idx, row in df_savings.iterrows():
        sku = str(row['SKU']).lower() if pd.notna(row['SKU']) else ''
        region = str(row['Region']).lower() if pd.notna(row['Region']) else ''
        os = str(row['OS']).lower() if pd.notna(row['OS']) else ''
        
        if not sku or not region or not os or os != 'windows':
            continue
        
        sku_region = f"{sku}_{region}"
        
        if sku_region in windows_pricing_data:
            pricing_info = windows_pricing_data[sku_region]
            description = str(row['Description']) if pd.notna(row['Description']) else ''
            
            if 'On Demand' in description or 'Pay-as-you-go' in description:
                cost = pricing_info.get('On Demand', None)
                if cost is not None:
                    df_savings.at[idx, 'Estimated monthly cost'] = f"${cost}" if not str(cost).startswith('$') else str(cost)
            elif '1 Year' in description:
                cost = pricing_info.get('1-Year Reserved', None)
                if cost is not None:
                    df_savings.at[idx, 'Estimated monthly cost'] = f"${cost}" if not str(cost).startswith('$') else str(cost)
            elif '3 Years' in description:
                cost = pricing_info.get('3-Year Reserved', None)
                if cost is not None:
                    df_savings.at[idx, 'Estimated monthly cost'] = f"${cost}" if not str(cost).startswith('$') else str(cost)
    
    # Recalculate totals for savings estimate
    df_savings['cost_numeric'] = df_savings.apply(lambda row: 
        float(str(row['Estimated monthly cost']).replace('$', '').replace(',', '')) 
        if pd.notna(row['Estimated monthly cost']) and row['Estimated monthly cost'] != '' and 'Total' not in str(row.get('Service category', ''))
        else 0, axis=1)
    
    total_payg_savings = df_savings[df_savings['Description'].str.contains('Pay-as-you-go', case=False, na=False)]['cost_numeric'].sum()
    total_1yr_savings = df_savings[df_savings['Description'].str.contains('1 Year', case=False, na=False) & ~df_savings['Description'].str.contains('Total', case=False, na=False)]['cost_numeric'].sum()
    total_3yr_savings = df_savings[df_savings['Description'].str.contains('3 Years', case=False, na=False) & ~df_savings['Description'].str.contains('Total', case=False, na=False)]['cost_numeric'].sum()
    
    # Update totals in savings dataframe
    annual_payg = total_payg_savings * 12
    annual_1yr = total_1yr_savings * 12
    annual_3yr = total_3yr_savings * 12
    savings_1yr = annual_payg - annual_1yr
    savings_3yr = annual_payg - annual_3yr
    total_36_months = total_payg_savings * 36
    total_3yr_cost = total_3yr_savings * 36
    
    # Update existing total rows and add savings rows
    mask_payg = df_savings['Description'] == 'Total Monthly Pay-as-you-go'
    mask_1yr = df_savings['Description'] == 'Total 1 Year Reservations (Billed monthly)'
    mask_3yr = df_savings['Description'] == 'Total 3 Year Reservations (Billed monthly)'
    
    df_savings.loc[mask_payg, 'Estimated monthly cost'] = f'${total_payg_savings:,.2f}'
    df_savings.loc[mask_1yr, 'Estimated monthly cost'] = f'${total_1yr_savings:,.2f}'
    df_savings.loc[mask_3yr, 'Estimated monthly cost'] = f'${total_3yr_savings:,.2f}'
    
    # Add additional savings rows with blank row formatting (3 lines + blank)
    blank_row = {'Service category': '', 'Service type': '', 'VM Name': '', 'Tags': '', 'Region': '', 
                 'OS': '', 'OS Name': '', 'SKU': '', 'Description': '', 'Estimated monthly cost': ''}
    
    additional_rows = [
        # Blank row after monthly totals
        blank_row.copy(),
        # Yearly and 1-year savings
        {'Service category': '', 'Service type': '', 'VM Name': '', 'Tags': '', 'Region': '', 
         'OS': '', 'OS Name': '', 'SKU': '', 'Description': 'Total Yearly Pay-as-you-go', 
         'Estimated monthly cost': f'${annual_payg:,.2f}'},
        {'Service category': '', 'Service type': '', 'VM Name': '', 'Tags': '', 'Region': '', 
         'OS': '', 'OS Name': '', 'SKU': '', 'Description': 'Total 1 Year Reservations (Annual cost)', 
         'Estimated monthly cost': f'${annual_1yr:,.2f}'},
        {'Service category': '', 'Service type': '', 'VM Name': '', 'Tags': '', 'Region': '', 
         'OS': '', 'OS Name': '', 'SKU': '', 'Description': 'Annual Savings (1 Year Reservations)', 
         'Estimated monthly cost': f'${savings_1yr:,.2f}'},
        # Blank row
        blank_row.copy(),
        # 3-year totals
        {'Service category': '', 'Service type': '', 'VM Name': '', 'Tags': '', 'Region': '', 
         'OS': '', 'OS Name': '', 'SKU': '', 'Description': 'Total 36 Months Pay-as-you-go', 
         'Estimated monthly cost': f'${total_36_months:,.2f}'},
        {'Service category': '', 'Service type': '', 'VM Name': '', 'Tags': '', 'Region': '', 
         'OS': '', 'OS Name': '', 'SKU': '', 'Description': 'Total 3 Year Reservations (3 Year cost)', 
         'Estimated monthly cost': f'${total_3yr_cost:,.2f}'},
        {'Service category': '', 'Service type': '', 'VM Name': '', 'Tags': '', 'Region': '', 
         'OS': '', 'OS Name': '', 'SKU': '', 'Description': 'Total 3 Year Reservations (annual cost)', 
         'Estimated monthly cost': f'${annual_3yr:,.2f}'},
        # Blank row
        blank_row.copy(),
        # Final savings row
        {'Service category': '', 'Service type': '', 'VM Name': '', 'Tags': '', 'Region': '', 
         'OS': '', 'OS Name': '', 'SKU': '', 'Description': '3 Year Reservations (Annual Savings)', 
         'Estimated monthly cost': f'${savings_3yr:,.2f}'},
    ]
    
    df_savings = df_savings.drop('cost_numeric', axis=1)
    df_savings = pd.concat([df_savings, pd.DataFrame(additional_rows)], ignore_index=True)
    
    # Build ranked_vms (just VM name and pay-as-you-go cost)
    ranked_data = []
    for vm_name in sorted_vm_names:
        ranked_data.append({
            'VM Name': vm_name,
            'Pay-as-you-go': f'${vm_pricing[vm_name]:,.2f}'
        })
    df_ranked = pd.DataFrame(ranked_data)
    
    return df_final, df_savings, df_ranked


def apply_excel_formatting(file_path):
    """Apply bold headers and borders to Excel file, and bold specific savings rows"""
    wb = load_workbook(file_path)
    ws = wb.active
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Bold and border the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.border = border_style
    
    # Bold the two annual savings rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Check the Description column (column I, index 8)
        description_cell = row[8]
        if description_cell.value in ['Annual Savings (1 Year Reservations)', '3 Year Reservations (Annual Savings)']:
            # Bold the Description cell and the Estimated monthly cost cell (column J, index 9)
            description_cell.font = Font(bold=True)
            row[9].font = Font(bold=True)
    
    wb.save(file_path)


def generate_pricing_spreadsheets():
    """Generate pricing spreadsheets and analysis"""
    print("\n" + "="*80)
    print("SECTION 2: GENERATING PRICING SPREADSHEETS")
    print("="*80 + "\n")
    
    # Load inputs
    INPUT_FILE = "inputs.json"
    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError(f"❌ Could not find {INPUT_FILE}")

    with open(INPUT_FILE, "r") as f:
        inputs = json.load(f)

    # Initialize Windows SKU-region tracking
    unique_skus_regions = set()
    
    # Save empty skus-regions-windows.json initially
    with open("skus-regions-windows.json", "w") as f:
        json.dump([], f, indent=4)
    print(f"✅ Created empty skus-regions-windows.json")

    print("\n=== Phase 1: Identifying Windows SKUs from inputs ===")
    # Quick pass to identify Windows SKUs for parallel scraping
    for row in inputs:
        if row.get("OS", "").lower() == "windows":
            sku = row["SKU"].replace("Standard_", "").lower().replace("_", "-")
            region = row["Region"].lower()
            unique_skus_regions.add(f"{sku}_{region}")
    
    print(f"Found {len(unique_skus_regions)} Windows SKU-region pairs to scrape")
    
    # Save skus-regions-windows.json for reference
    with open("skus-regions-windows.json", "w") as f:
        json.dump(list(unique_skus_regions), f, indent=4)

    # Start both operations in parallel using threads
    print("\n=== Phase 2: Running Azure API and Selenium scraping in parallel ===")
    
    estimate_rows = []
    windows_pricing_data = {}
    exceptions = []
    
    def azure_api_thread():
        nonlocal estimate_rows
        try:
            print("Azure API thread started...")
            estimate_rows, _ = build_azure_pricing(inputs)
            print("Azure API thread completed!")
        except Exception as e:
            exceptions.append(("Azure API", e))
    
    def selenium_thread():
        nonlocal windows_pricing_data
        try:
            print("Selenium scraping thread started...")
            windows_pricing_data = scrape_windows_pricing(unique_skus_regions)
            print("Selenium scraping thread completed!")
        except Exception as e:
            exceptions.append(("Selenium", e))
    
    # Create and start threads
    azure_thread = threading.Thread(target=azure_api_thread)
    selenium_thread = threading.Thread(target=selenium_thread)
    
    azure_thread.start()
    selenium_thread.start()
    
    # Wait for both to complete
    azure_thread.join()
    selenium_thread.join()
    
    # Check for exceptions
    if exceptions:
        for name, exc in exceptions:
            print(f"❌ Error in {name}: {exc}")
        raise RuntimeError("One or more threads failed")
    
    print(f"\n✅ Both operations completed successfully!")
    
    # Save Windows pricing data
    with open('azure_windows_pricing_data.json', 'w') as f:
        json.dump(windows_pricing_data, f, indent=4)
    print(f"✅ Saved azure_windows_pricing_data.json")

    print("\n=== Building final spreadsheets ===")
    df_estimate, df_savings, df_ranked = build_final_dataframes(estimate_rows, windows_pricing_data)
    
    # Save Excel files
    df_estimate.to_excel("azure_estimate.xlsx", index=False)
    print(f"✅ Created azure_estimate.xlsx with {len(df_estimate)} rows")
    
    # For azure_savings_estimate, create with FilterMe sheet
    with pd.ExcelWriter("azure_savings_estimate.xlsx", engine='openpyxl') as writer:
        df_savings.to_excel(writer, index=False, sheet_name='Sheet1')
        # Create FilterMe sheet (flattened version without blank rows)
        # Remove rows where Description is empty (blank rows)
        df_savings_flat = df_savings[
            (df_savings['Description'].notna()) & 
            (df_savings['Description'] != '')
        ].copy()
        df_savings_flat.to_excel(writer, index=False, sheet_name='FilterMe')
    print(f"✅ Created azure_savings_estimate.xlsx with {len(df_savings)} rows")
    
    df_ranked.to_excel("ranked_vms.xlsx", index=False)
    print(f"✅ Created ranked_vms.xlsx with {len(df_ranked)} rows")
    
    # Apply formatting
    apply_excel_formatting("azure_savings_estimate.xlsx")
    print(f"✅ Applied formatting to azure_savings_estimate.xlsx")


###############################################################################
# MAIN PROGRAM
###############################################################################

def main():
    """Main program - runs both sections with user prompt between"""
    print("\n" + "="*80)
    print("AZURE RESERVATION ANALYSIS TOOL")
    print("="*80)
    print("\n⚠️ NOTE: This script requires Chrome and ChromeDriver to be installed")
    print("   for web scraping. If you encounter errors in Section 2, please install:")
    print("   - Google Chrome browser")
    print("   - ChromeDriver (matching your Chrome version)")
    print("   Download: https://chromedriver.chromium.org/\n")
    
    try:
        # Section 1: Generate VM recommendations
        generate_vm_recommendations()
        
        # Prompt user to continue
        print("\n" + "="*80)
        user_input = input("Continue to pricing analysis and spreadsheet generation? (yes/no): ").strip().lower()
        
        if user_input in ['yes', 'y']:
            # Section 2: Generate pricing spreadsheets
            generate_pricing_spreadsheets()
            
            print("\n" + "="*80)
            print("✅ ALL PROCESSING COMPLETE!")
            print("="*80)
            print("\nGenerated Files:")
            print("  - output.json (VM recommendations)")
            print("  - inputs.json (Input for pricing)")
            print("  - azure_estimate.xlsx (Detailed pricing)")
            print("  - azure_savings_estimate.xlsx (With vantage.sh pricing)")
            print("  - ranked_vms.xlsx (Sorted by cost)")
            print("  - azure_windows_pricing_data.json (Windows pricing cache)")
            print("  - skus-regions-windows.json (SKU-region pairs)")
        else:
            print("\n⏸️ Stopped after recommendations generation.")
            print("Generated Files:")
            print("  - output.json (VM recommendations)")
            print("  - inputs.json (Input for pricing)")
            
    except Exception as e:
        print(f"\n❌ Error: {e}")
        raise


if __name__ == "__main__":
    main()

